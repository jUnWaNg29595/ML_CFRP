from __future__ import annotations

import argparse
import csv
import json
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from repair_gemini_bigsmiles_cache import (
    is_bigsmiles_candidate,
    repair_bigsmiles_text,
    split_top_level,
    validate_bigsmiles_text,
    validate_smiles_text,
)


ROOT_DIR = Path(__file__).resolve().parent
SOURCE_CANDIDATES = [
    ROOT_DIR / "备份数据" / "环氧数据_纯数据.xlsx",
    ROOT_DIR / "source_epoxy_data.xlsx",
]
CANONICAL_EXPORT = ROOT_DIR / "备份数据" / "data_export.csv"
CANONICAL_REPORT = ROOT_DIR / "chemical_parse_report.csv"
REPAIR_SCRIPT = ROOT_DIR / "repair_chemical_parse_bigsmiles_with_gemini.py"

RAW_COLUMNS = ("resin_smiles", "curing_agent_smiles")
MIN_RESIN_COMPONENT_COLUMNS = 6
MIN_CURING_COMPONENT_COLUMNS = 10

FULLWIDTH_TRANSLATION = str.maketrans(
    {
        "；": ";",
        "（": "(",
        "）": ")",
        "，": ",",
        "：": ":",
        "。": ".",
        "\u3000": " ",
    }
)

ALTERNATIVE_PHRASES = ("或", " or ", " OR ")
GROUP_PHRASES = ("和", "与", "及", " and ", " AND ")

ANNOTATION_PATTERNS = [
    re.compile(r"G\d+\s*代表性SMILES", re.IGNORECASE),
    re.compile(r"代表性单体SMILES", re.IGNORECASE),
    re.compile(r"代表性环氧树脂", re.IGNORECASE),
    re.compile(r"代表性SMILES", re.IGNORECASE),
    re.compile(r"BigSMILES", re.IGNORECASE),
    re.compile(r"\brepresentative(?:\s+monomer)?\s+SMILES\b", re.IGNORECASE),
    re.compile(r"\bSMILES\b", re.IGNORECASE),
    re.compile(r"无法表示", re.IGNORECASE),
    re.compile(r"不能表示", re.IGNORECASE),
    re.compile(r"unrepresentable", re.IGNORECASE),
    re.compile(r"not representable", re.IGNORECASE),
    re.compile(r"cannot be represented", re.IGNORECASE),
]

PURE_NOISE_PATTERNS = [
    re.compile(r"^(?:无法表示|不能表示)$", re.IGNORECASE),
    re.compile(r"^(?:unrepresentable|not representable|cannot be represented)$", re.IGNORECASE),
    re.compile(r"^(?:bigsmiles|smiles)$", re.IGNORECASE),
]

NOTE_KEYWORDS = (
    "bigsmiles",
    "smiles",
    "representative",
    "无法表示",
    "不能表示",
    "聚苯乙烯段",
    "聚丁二烯段",
    "十二烷基",
    "十四烷基",
)


@dataclass
class CellCleanResult:
    original: str
    cleaned_text: str
    groups: list[list[str]]
    notes: list[str]

    @property
    def components(self) -> list[str]:
        return [component for group in self.groups for component in group]


def resolve_default_source() -> Path:
    for candidate in SOURCE_CANDIDATES:
        if candidate.exists():
            return candidate
    return SOURCE_CANDIDATES[0]


def backup_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.name}.bak_{stamp}")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def normalize_text(text: str) -> str:
    updated = str(text or "").translate(FULLWIDTH_TRANSLATION)
    updated = updated.replace("\r", " ").replace("\n", " ")
    updated = re.sub(r"\s+", " ", updated)
    updated = re.sub(r"\s*;\s*", "; ", updated)
    updated = re.sub(r"\s*\.\s*", ".", updated)
    return updated.strip(" ;")


def is_note_parenthetical(body: str) -> bool:
    stripped = body.strip()
    if not stripped:
        return True
    lowered = stripped.lower()
    if any(keyword in lowered for keyword in NOTE_KEYWORDS):
        return True
    if re.search(r"[\u4e00-\u9fff]", stripped):
        return True
    if re.fullmatch(r"[A-Za-z0-9 _+\-]{1,40}", stripped):
        return True
    return False


def strip_note_parentheticals(text: str, notes: list[str]) -> str:
    pattern = re.compile(r"\s+[（(](?P<body>[^()（）]{1,120})[)）]")
    updated = text
    while True:
        match = pattern.search(updated)
        if match is None:
            break
        body = match.group("body")
        if not is_note_parenthetical(body):
            break
        notes.append(f"drop_parenthetical:{body.strip()}")
        updated = updated[: match.start()] + updated[match.end() :]
    return updated


def remove_annotation_words(text: str, notes: list[str]) -> str:
    updated = text
    for pattern in ANNOTATION_PATTERNS:
        while True:
            match = pattern.search(updated)
            if match is None:
                break
            notes.append(f"drop_annotation:{match.group(0)}")
            updated = updated[: match.start()] + updated[match.end() :]
    updated = re.sub(r"\s+", " ", updated)
    updated = re.sub(r"\.{2,}", ".", updated)
    updated = re.sub(r";{2,}", ";", updated)
    return updated.strip(" ;.")


def split_top_level_phrases(text: str, phrases: tuple[str, ...]) -> list[str]:
    parts: list[str] = []
    current: list[str] = []
    paren_depth = 0
    bracket_depth = 0
    brace_depth = 0
    index = 0
    ordered_phrases = sorted(phrases, key=len, reverse=True)

    while index < len(text):
        if paren_depth == 0 and bracket_depth == 0 and brace_depth == 0:
            matched = next((phrase for phrase in ordered_phrases if text.startswith(phrase, index)), None)
            if matched is not None:
                part = "".join(current).strip()
                if part:
                    parts.append(part)
                current = []
                index += len(matched)
                continue

        ch = text[index]
        if ch == "[":
            bracket_depth += 1
        elif ch == "]":
            bracket_depth = max(0, bracket_depth - 1)
        elif ch == "(":
            paren_depth += 1
        elif ch == ")":
            paren_depth = max(0, paren_depth - 1)
        elif ch == "{":
            brace_depth += 1
        elif ch == "}":
            brace_depth = max(0, brace_depth - 1)

        current.append(ch)
        index += 1

    tail = "".join(current).strip()
    if tail:
        parts.append(tail)
    return parts


def clean_candidate_preview(text: str) -> str:
    scratch: list[str] = []
    updated = normalize_text(text)
    updated = strip_note_parentheticals(updated, scratch)
    updated = remove_annotation_words(updated, scratch)
    return updated.strip(" ;.")


def candidate_score(candidate: str, index: int) -> tuple[int, int, int, int, int, int]:
    raw = str(candidate or "").strip()
    preview = clean_candidate_preview(raw)
    explicit_bigsmiles = int("bigsmiles" in raw.lower())
    representative_penalty = int(
        "代表性" in raw or "representative" in raw.lower()
    )
    bigsmiles_like = int(is_bigsmiles_candidate(preview))
    bigsmiles_valid = 0
    smiles_valid = 0
    if preview:
        if bigsmiles_like:
            bigsmiles_valid = int(repair_bigsmiles_text(preview).valid_after)
        else:
            smiles_valid = int(validate_smiles_text(preview).get("valid"))
    structural_weight = sum(ch.isalnum() or ch in "[]{}()=#@+-%\\/." for ch in preview)
    return (
        explicit_bigsmiles,
        bigsmiles_like,
        bigsmiles_valid,
        1 - representative_penalty,
        smiles_valid,
        structural_weight - index,
    )


def choose_preferred_candidate(candidates: list[str]) -> str:
    scored = sorted(
        ((candidate_score(candidate, index), candidate) for index, candidate in enumerate(candidates)),
        key=lambda item: item[0],
        reverse=True,
    )
    return scored[0][1] if scored else ""


def is_pure_noise(text: str) -> bool:
    stripped = str(text or "").strip()
    if not stripped:
        return True
    return any(pattern.fullmatch(stripped) for pattern in PURE_NOISE_PATTERNS)


def has_structure_signal(text: str) -> bool:
    return bool(re.search(r"[A-Za-z0-9\[\]\(\)\{\}=#@+\-\\/:%]", text))


def clean_component_text(text: str, notes: list[str]) -> str:
    updated = normalize_text(text)
    updated = strip_note_parentheticals(updated, notes)
    updated = remove_annotation_words(updated, notes)
    updated = normalize_text(updated)
    updated = updated.strip(" ;.")
    if not updated:
        return ""
    if is_pure_noise(updated):
        notes.append(f"drop_noise_component:{text.strip()}")
        return ""
    if not has_structure_signal(updated):
        notes.append(f"drop_non_structural_component:{text.strip()}")
        return ""
    if is_bigsmiles_candidate(updated):
        repair = repair_bigsmiles_text(updated)
        if repair.valid_after and repair.repaired != updated:
            notes.append(f"local_bigsmiles_repair:{updated}->{repair.repaired}")
            updated = repair.repaired
    return updated


def clean_smiles_cell(value: Any) -> CellCleanResult:
    original = str(value or "").strip()
    if not original:
        return CellCleanResult(original="", cleaned_text="", groups=[], notes=[])

    notes: list[str] = []
    working = normalize_text(original)
    working = strip_note_parentheticals(working, notes)

    coarse_groups: list[str] = []
    for segment in split_top_level(working, {";"}):
        segment = segment.strip()
        if not segment:
            continue
        coarse_groups.extend(split_top_level_phrases(segment, GROUP_PHRASES))

    cleaned_groups: list[list[str]] = []
    for group_text in coarse_groups:
        alternatives = split_top_level_phrases(group_text, ALTERNATIVE_PHRASES)
        chosen_group = group_text
        if len(alternatives) > 1:
            chosen_group = choose_preferred_candidate(alternatives)
            notes.append(f"selected_alternative:{clean_candidate_preview(chosen_group)}")
            for alternative in alternatives:
                if alternative != chosen_group:
                    notes.append(f"dropped_alternative:{clean_candidate_preview(alternative)}")

        group_cleaned = strip_note_parentheticals(chosen_group, notes)
        group_cleaned = remove_annotation_words(group_cleaned, notes)
        group_cleaned = normalize_text(group_cleaned)
        raw_components = split_top_level(group_cleaned, {"."})

        cleaned_components: list[str] = []
        for raw_component in raw_components:
            component = clean_component_text(raw_component, notes)
            if component:
                cleaned_components.append(component)

        if cleaned_components:
            cleaned_groups.append(cleaned_components)

    cleaned_text = "; ".join(".".join(group) for group in cleaned_groups)
    if not cleaned_text:
        cleaned_text = clean_candidate_preview(working)
    deduped_notes = list(dict.fromkeys(note for note in notes if note))
    return CellCleanResult(
        original=original,
        cleaned_text=cleaned_text,
        groups=cleaned_groups,
        notes=deduped_notes,
    )


def build_export_headers(
    source_headers: list[str],
    max_resin_components: int,
    max_curing_components: int,
) -> list[str]:
    base_headers = [header for header in source_headers if header not in RAW_COLUMNS]
    resin_headers = [f"resin_smiles_{index}" for index in range(1, max_resin_components + 1)]
    curing_headers = [f"curing_agent_smiles_{index}" for index in range(1, max_curing_components + 1)]
    return (
        base_headers
        + resin_headers
        + ["resin_smiles_n_components"]
        + curing_headers
        + ["curing_agent_smiles_n_components"]
    )


def make_parse_report_row(row_index: int, column: str, chemical_text: str) -> dict[str, Any]:
    text = str(chemical_text or "").strip()
    if not text:
        return {}

    if is_bigsmiles_candidate(text):
        direct = validate_bigsmiles_text(text)
        fallback = repair_bigsmiles_text(text)
        fallback_smiles_validation = validate_smiles_text(fallback.fallback_smiles) if fallback.fallback_smiles else {
            "valid": False,
            "canonical": "",
        }
        if direct.get("valid"):
            return {
                "row_index": row_index,
                "column": column,
                "chemical_raw": text,
                "format": "bigsmiles",
                "status": "ok",
                "partial_parse": True,
                "rdkit_direct_ok": False,
                "proxy_smiles_ok": False,
                "normalized_ok": True,
                "proxy_smiles": "",
                "normalized_smiles": text,
                "invalid_fragments": "",
                "valid_fragments": text,
                "reason": "bigsmiles_direct",
                "details": "BigSMILES validator accepted the cleaned notation.",
            }

        details_bits = [str(direct.get("error", "") or "").strip()]
        if fallback.applied_rules:
            details_bits.append("local_rules=" + "|".join(fallback.applied_rules))
        if fallback.repaired and fallback.repaired != text:
            details_bits.append("local_candidate=" + fallback.repaired)
        if fallback_smiles_validation.get("valid"):
            details_bits.append("fallback_smiles=" + str(fallback_smiles_validation.get("canonical", "")))

        return {
            "row_index": row_index,
            "column": column,
            "chemical_raw": text,
            "format": "bigsmiles",
            "status": "invalid",
            "partial_parse": bool(fallback_smiles_validation.get("valid")),
            "rdkit_direct_ok": False,
            "proxy_smiles_ok": bool(fallback_smiles_validation.get("valid")),
            "normalized_ok": False,
            "proxy_smiles": str(fallback_smiles_validation.get("canonical", "")),
            "normalized_smiles": fallback.repaired if fallback.repaired != text else "",
            "invalid_fragments": text,
            "valid_fragments": str(fallback_smiles_validation.get("canonical", "")),
            "reason": "parse_failed",
            "details": " | ".join(bit for bit in details_bits if bit),
        }

    smiles_validation = validate_smiles_text(text)
    if smiles_validation.get("valid"):
        canonical = str(smiles_validation.get("canonical", "") or text)
        return {
            "row_index": row_index,
            "column": column,
            "chemical_raw": text,
            "format": "smiles",
            "status": "ok",
            "partial_parse": True,
            "rdkit_direct_ok": True,
            "proxy_smiles_ok": True,
            "normalized_ok": True,
            "proxy_smiles": canonical,
            "normalized_smiles": canonical,
            "invalid_fragments": "",
            "valid_fragments": canonical,
            "reason": "rdkit_direct",
            "details": "",
        }

    return {
        "row_index": row_index,
        "column": column,
        "chemical_raw": text,
        "format": "smiles",
        "status": "invalid",
        "partial_parse": False,
        "rdkit_direct_ok": False,
        "proxy_smiles_ok": False,
        "normalized_ok": False,
        "proxy_smiles": "",
        "normalized_smiles": "",
        "invalid_fragments": text,
        "valid_fragments": "",
        "reason": "parse_failed",
        "details": str(smiles_validation.get("error", "") or "RDKit could not parse the cleaned SMILES."),
    }


def apply_cleaned_values_to_workbook(
    source_path: Path,
    workbook_out: Path,
    sheet_name: str,
    row_results: dict[int, dict[str, CellCleanResult]],
    header_map: dict[str, int],
) -> None:
    wb = load_workbook(source_path)
    ws = wb[sheet_name]

    extra_headers = [
        "resin_smiles_original_text",
        "resin_smiles_clean_notes",
        "curing_agent_smiles_original_text",
        "curing_agent_smiles_clean_notes",
    ]
    extra_column_map: dict[str, int] = {}
    next_column = ws.max_column + 1
    for header in extra_headers:
        extra_column_map[header] = next_column
        ws.cell(row=1, column=next_column).value = header
        next_column += 1

    for row_index, per_column in row_results.items():
        resin_result = per_column["resin_smiles"]
        curing_result = per_column["curing_agent_smiles"]

        ws.cell(row=row_index, column=header_map["resin_smiles"]).value = resin_result.cleaned_text
        ws.cell(row=row_index, column=header_map["curing_agent_smiles"]).value = curing_result.cleaned_text

        ws.cell(row=row_index, column=extra_column_map["resin_smiles_original_text"]).value = (
            resin_result.original if resin_result.cleaned_text != resin_result.original else ""
        )
        ws.cell(row=row_index, column=extra_column_map["resin_smiles_clean_notes"]).value = " | ".join(
            resin_result.notes
        )
        ws.cell(row=row_index, column=extra_column_map["curing_agent_smiles_original_text"]).value = (
            curing_result.original if curing_result.cleaned_text != curing_result.original else ""
        )
        ws.cell(row=row_index, column=extra_column_map["curing_agent_smiles_clean_notes"]).value = " | ".join(
            curing_result.notes
        )

    workbook_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_out)


def maybe_copy_with_backup(source_path: Path, destination_path: Path) -> Path | None:
    if not source_path.exists():
        return None
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    if destination_path.exists():
        backup = backup_path(destination_path)
        shutil.copy2(destination_path, backup)
    shutil.copy2(source_path, destination_path)
    return destination_path


def run_bigsmiles_repair(
    python_executable: str,
    report_path: Path,
    mapping_out: Path,
    rows_out: Path,
    annotated_report_out: Path,
    *,
    statuses: str,
    limit: int,
    no_ai: bool,
) -> None:
    if not REPAIR_SCRIPT.exists():
        raise FileNotFoundError(REPAIR_SCRIPT)

    command = [
        python_executable,
        str(REPAIR_SCRIPT),
        "--report",
        str(report_path),
        "--statuses",
        statuses,
        "--limit",
        str(limit),
        "--mapping-out",
        str(mapping_out),
        "--rows-out",
        str(rows_out),
        "--annotated-report-out",
        str(annotated_report_out),
        "--no-gui",
    ]
    if no_ai:
        command.append("--no-ai")
    subprocess.run(command, cwd=str(ROOT_DIR), check=True)


def build_parser() -> argparse.ArgumentParser:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_source = resolve_default_source()
    default_workbook_out = ROOT_DIR / "备份数据" / f"环氧数据_纯数据_smiles_cleaned_{stamp}.xlsx"
    default_export_out = ROOT_DIR / "备份数据" / f"data_export_cleaned_from_source_{stamp}.csv"
    default_report_out = ROOT_DIR / f"chemical_parse_report_from_source_{stamp}.csv"
    default_audit_out = ROOT_DIR / f"source_smiles_cleanup_audit_{stamp}.csv"
    default_repair_map = ROOT_DIR / f"chemical_parse_bigsmiles_ai_repair_map_source_{stamp}.csv"
    default_repair_rows = ROOT_DIR / f"chemical_parse_bigsmiles_ai_repair_rows_source_{stamp}.csv"
    default_repair_annotated = ROOT_DIR / f"chemical_parse_report_with_ai_repair_preview_source_{stamp}.csv"

    parser = argparse.ArgumentParser(
        description=(
            "Clean AI-polluted SMILES text in the source epoxy workbook, regenerate "
            "component-split CSV output, rebuild chemical_parse_report.csv, and optionally "
            "rerun the BIGSMILES repair pipeline."
        )
    )
    parser.add_argument("--source", default=str(default_source), help="Input workbook path.")
    parser.add_argument("--sheet", default="Sheet1", help="Worksheet name.")
    parser.add_argument("--workbook-out", default=str(default_workbook_out), help="Cleaned workbook copy.")
    parser.add_argument("--export-out", default=str(default_export_out), help="Regenerated split export CSV.")
    parser.add_argument("--report-out", default=str(default_report_out), help="Regenerated parse report CSV.")
    parser.add_argument("--audit-out", default=str(default_audit_out), help="Audit CSV for changed source cells.")
    parser.add_argument(
        "--overwrite-canonical",
        action="store_true",
        help="Also overwrite canonical outputs at 备份数据/data_export.csv and chemical_parse_report.csv with backups.",
    )
    parser.add_argument(
        "--run-bigsmiles-repair",
        action="store_true",
        help="Invoke repair_chemical_parse_bigsmiles_with_gemini.py on the regenerated parse report.",
    )
    parser.add_argument(
        "--repair-no-ai",
        action="store_true",
        help="When rerunning BIGSMILES repair, only use local rules.",
    )
    parser.add_argument(
        "--repair-statuses",
        default="invalid",
        help="Comma-separated statuses for the BIGSMILES repair rerun.",
    )
    parser.add_argument(
        "--repair-limit",
        type=int,
        default=0,
        help="Optional limit for unique BIGSMILES cases during the repair rerun.",
    )
    parser.add_argument("--repair-mapping-out", default=str(default_repair_map), help="BIGSMILES repair mapping CSV.")
    parser.add_argument("--repair-rows-out", default=str(default_repair_rows), help="BIGSMILES repair per-row CSV.")
    parser.add_argument(
        "--repair-annotated-report-out",
        default=str(default_repair_annotated),
        help="BIGSMILES repair annotated parse report.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()

    source_path = Path(args.source)
    workbook_out = Path(args.workbook_out)
    export_out = Path(args.export_out)
    report_out = Path(args.report_out)
    audit_out = Path(args.audit_out)

    if not source_path.exists():
        raise FileNotFoundError(source_path)

    wb = load_workbook(source_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Worksheet not found: {args.sheet}")
    ws = wb[args.sheet]

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    source_headers = [str(header or "").strip() for header in header_row]
    header_map = {header: index for index, header in enumerate(source_headers, start=1)}
    for required_header in RAW_COLUMNS:
        if required_header not in header_map:
            raise KeyError(f"Missing required column: {required_header}")

    row_results: dict[int, dict[str, CellCleanResult]] = {}
    export_rows: list[dict[str, Any]] = []
    parse_report_rows: list[dict[str, Any]] = []
    audit_rows: list[dict[str, Any]] = []

    max_resin_components = MIN_RESIN_COMPONENT_COLUMNS
    max_curing_components = MIN_CURING_COMPONENT_COLUMNS
    changed_cells = 0
    nonempty_source_cells = 0
    chinese_source_cells = 0

    for row_index, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {
            source_headers[index]: value
            for index, value in enumerate(row_values)
            if index < len(source_headers)
        }
        resin_result = clean_smiles_cell(row_dict.get("resin_smiles", ""))
        curing_result = clean_smiles_cell(row_dict.get("curing_agent_smiles", ""))
        row_results[row_index] = {
            "resin_smiles": resin_result,
            "curing_agent_smiles": curing_result,
        }

        for column_name, result in (("resin_smiles", resin_result), ("curing_agent_smiles", curing_result)):
            if result.original:
                nonempty_source_cells += 1
            if re.search(r"[\u4e00-\u9fff]", result.original):
                chinese_source_cells += 1
            if result.cleaned_text != result.original or result.notes:
                changed_cells += 1
                audit_rows.append(
                    {
                        "row_index": row_index,
                        "column": column_name,
                        "original_text": result.original,
                        "cleaned_text": result.cleaned_text,
                        "group_count": len(result.groups),
                        "component_count": len(result.components),
                        "notes": " | ".join(result.notes),
                    }
                )

        max_resin_components = max(max_resin_components, len(resin_result.components))
        max_curing_components = max(max_curing_components, len(curing_result.components))

    export_headers = build_export_headers(source_headers, max_resin_components, max_curing_components)
    resin_component_headers = [f"resin_smiles_{index}" for index in range(1, max_resin_components + 1)]
    curing_component_headers = [f"curing_agent_smiles_{index}" for index in range(1, max_curing_components + 1)]

    for row_index, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {
            source_headers[index]: row_values[index]
            for index in range(min(len(source_headers), len(row_values)))
        }
        resin_result = row_results[row_index]["resin_smiles"]
        curing_result = row_results[row_index]["curing_agent_smiles"]

        export_row = {
            header: row_dict.get(header, "")
            for header in source_headers
            if header not in RAW_COLUMNS
        }
        for index, header in enumerate(resin_component_headers):
            export_row[header] = resin_result.components[index] if index < len(resin_result.components) else ""
        export_row["resin_smiles_n_components"] = len(resin_result.components)
        for index, header in enumerate(curing_component_headers):
            export_row[header] = curing_result.components[index] if index < len(curing_result.components) else ""
        export_row["curing_agent_smiles_n_components"] = len(curing_result.components)
        export_rows.append(export_row)

        for header in resin_component_headers:
            component_text = export_row.get(header, "")
            if component_text:
                parse_report_rows.append(make_parse_report_row(row_index, header, component_text))
        for header in curing_component_headers:
            component_text = export_row.get(header, "")
            if component_text:
                parse_report_rows.append(make_parse_report_row(row_index, header, component_text))

    report_rows = [row for row in parse_report_rows if row]
    report_fields = [
        "row_index",
        "column",
        "chemical_raw",
        "format",
        "status",
        "partial_parse",
        "rdkit_direct_ok",
        "proxy_smiles_ok",
        "normalized_ok",
        "proxy_smiles",
        "normalized_smiles",
        "invalid_fragments",
        "valid_fragments",
        "reason",
        "details",
    ]
    audit_fields = [
        "row_index",
        "column",
        "original_text",
        "cleaned_text",
        "group_count",
        "component_count",
        "notes",
    ]

    write_csv(export_out, export_rows, export_headers)
    write_csv(report_out, report_rows, report_fields)
    write_csv(audit_out, audit_rows, audit_fields)
    apply_cleaned_values_to_workbook(
        source_path=source_path,
        workbook_out=workbook_out,
        sheet_name=args.sheet,
        row_results=row_results,
        header_map=header_map,
    )

    if args.overwrite_canonical:
        maybe_copy_with_backup(export_out, CANONICAL_EXPORT)
        maybe_copy_with_backup(report_out, CANONICAL_REPORT)

    repair_outputs: dict[str, str] = {}
    if args.run_bigsmiles_repair:
        target_report = CANONICAL_REPORT if args.overwrite_canonical else report_out
        run_bigsmiles_repair(
            python_executable=sys.executable,
            report_path=target_report,
            mapping_out=Path(args.repair_mapping_out),
            rows_out=Path(args.repair_rows_out),
            annotated_report_out=Path(args.repair_annotated_report_out),
            statuses=args.repair_statuses,
            limit=args.repair_limit,
            no_ai=args.repair_no_ai,
        )
        repair_outputs = {
            "repair_mapping_out": str(Path(args.repair_mapping_out)),
            "repair_rows_out": str(Path(args.repair_rows_out)),
            "repair_annotated_report_out": str(Path(args.repair_annotated_report_out)),
            "repair_used_ai": str(not args.repair_no_ai),
        }

    status_counts: dict[str, int] = {}
    format_counts: dict[str, int] = {}
    for row in report_rows:
        status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1
        format_counts[row["format"]] = format_counts.get(row["format"], 0) + 1

    summary = {
        "source": str(source_path),
        "sheet": args.sheet,
        "workbook_out": str(workbook_out),
        "export_out": str(export_out),
        "report_out": str(report_out),
        "audit_out": str(audit_out),
        "overwrite_canonical": bool(args.overwrite_canonical),
        "source_nonempty_smiles_cells": nonempty_source_cells,
        "source_cells_with_chinese": chinese_source_cells,
        "changed_smiles_cells": changed_cells,
        "rows_processed": len(export_rows),
        "report_rows": len(report_rows),
        "max_resin_components": max_resin_components,
        "max_curing_agent_components": max_curing_components,
        "report_status_counts": status_counts,
        "report_format_counts": format_counts,
        **repair_outputs,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
